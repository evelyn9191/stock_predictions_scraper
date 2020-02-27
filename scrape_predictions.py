import logging
import random
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from random import randint
from shutil import copyfile
from typing import List

import openpyxl as openpyxl
import requests

from bs4 import BeautifulSoup
from openpyxl import Workbook

from user_agents import user_agent_list

logging.getLogger().setLevel(logging.INFO)


def check_latest_predictions(file_with_predictions: str) -> bool:
    if not Path(file_with_predictions).exists():
        return True

    today = date.today()

    wb = openpyxl.load_workbook(file_with_predictions)
    latest_predictions = datetime.strptime(wb.sheetnames[-1], "%d-%m-%Y").date()
    logging.debug(f"Latest predictions: {latest_predictions}")

    if latest_predictions and today - latest_predictions <= timedelta(days=29):
        return False

    return True


def get_stock_identifiers(file_with_stocks: str):
    wb = openpyxl.load_workbook(file_with_stocks)
    ws = wb.active

    stocks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            stocks.append(row)

    logging.debug(stocks)
    return stocks


def get_webpage(stock_code: str) -> str:
    url = f"https://walletinvestor.com/stock-forecast?currency={stock_code}"
    user_agent = random.choice(user_agent_list)
    headers = {"User-Agent": user_agent}

    response = requests.get(url, headers=headers)
    html = response.text
    return html


def parse_table(webpage: str) -> list:
    stock_predictions = []

    soup = BeautifulSoup(webpage, "lxml")

    rating = soup.select_one("[class~=currency-rate]").string
    logging.debug(f"Rating: {rating}")
    stock_predictions.append(rating)

    price = soup.find(class_="kv-align-right kv-align-middle w3").string
    logging.debug(f"Price: {price}")
    stock_predictions.append(price)

    prediction_intervals = {"14d": 4, "3m": 5, "6m": 6, "1y": 7, "5y": 8}

    for _, column_sequence in prediction_intervals.items():
        prediction = _get_performance_prediction(soup, column_sequence)
        stock_predictions.append(prediction)

    logging.info(stock_predictions)
    return stock_predictions


def _get_performance_prediction(soup: BeautifulSoup, column_number: int) -> float:
    prediction = soup.find(
        "td",
        attrs={
            "class": "table-cell-label kv-align-right kv-align-middle w3",
            "data-col-seq": column_number,
        },
    ).get_text()

    clean_prediction = prediction.replace(" ", "").strip("+ %")

    return float(clean_prediction) / 100


def add_stock_identification(stock_identifiers: tuple, stock_data: list) -> list:
    stock_row = list(stock_identifiers) + stock_data
    logging.debug(stock_row)
    return stock_row


def write_to_excel(
    source_file: str, file_to_write: str, predictions: List[list], sheetname: str
) -> Workbook:
    if Path(file_to_write).exists():
        wb = openpyxl.load_workbook(file_to_write)
    else:
        copyfile(source_file, file_to_write)
        wb = openpyxl.load_workbook(file_to_write)

    ws = wb.create_sheet(title=sheetname)

    column_headers = [
        "Company",
        "Stock Code",
        "Rating",
        "Price (USD)",
        "14d prediction",
        "3m prediction",
        "6m prediction",
        "1y prediction",
        "5y prediction",
    ]
    ws.append(column_headers)

    for one_stock in predictions:
        ws.append(one_stock)

    wb.save(file_to_write)

    return wb


def get_stock_predictions(file_with_stocks: str) -> list:
    predictions = []
    stocks_to_predict = get_stock_identifiers(file_with_stocks)

    for stock in stocks_to_predict:
        page = get_webpage(stock[1])
        try:
            parsed_stock_values = parse_table(page)
        except AttributeError:
            continue
        stock_row = add_stock_identification(stock, parsed_stock_values)
        predictions.append(stock_row)
        time.sleep(randint(3, 10))

    sorted(predictions, key=lambda x: x[1])

    return predictions
