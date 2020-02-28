import logging
import random
import time
from random import randint

import openpyxl as openpyxl

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.webdriver import WebDriver

from scraper.user_agents import user_agent_list

logging.getLogger().setLevel(logging.INFO)


def initialize_driver() -> WebDriver:
    chrome_options = Options()
    user_agent = random.choice(user_agent_list)
    chrome_options.add_argument(f"user-agent=[{user_agent}]")
    driver = webdriver.Chrome("../chromedrivers/win.exe", chrome_options=chrome_options)
    return driver


def get_tickers(file_with_stocks: str) -> list:
    wb = openpyxl.load_workbook(file_with_stocks)
    ws = wb.active

    tickers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tickers.append(row[1])

    logging.debug(tickers)
    return tickers


def get_consensus(driver: WebDriver, ticker: str) -> str:
    driver.get(f"https://www.nasdaq.com/market-activity/stocks/{ticker}/analyst-research")
    consensus_xpath = (
        "/html/body/div[4]/div/main/div/div[4]/div[2]/div/div[1]/div[2]/div[2]/div/span[2]"
    )
    ActionChains(driver).move_to_element(driver.find_element_by_xpath(consensus_xpath)).perform()
    consensus = driver.find_element_by_xpath(consensus_xpath).text
    print(f"{ticker + ': ' + consensus}")
    return consensus


def add_consensus_column(file_with_predictions: str, sheetname: str, consensuses: list) -> None:
    wb = openpyxl.load_workbook(file_with_predictions)
    ws = wb[sheetname]

    consensuses.insert(0, ["J1", "Consensus"])

    for consensus in consensuses:
        cell = ws[consensus[0]]
        cell.value = consensus[1]

    wb.save(file_with_predictions)


def add_consensus(file_with_predictions: str, sheet_name: str) -> None:
    driver = initialize_driver()
    tickers = get_tickers(file_with_predictions)
    ticker_consensuses = []

    logging.info("Scraping consensuses...")
    current_index = 2
    unknown_consensus = "N/A"
    for ticker in tickers:
        try:
            analysts_consensus = get_consensus(driver, ticker)
        except NoSuchElementException:
            analysts_consensus = unknown_consensus
        ticker_consensuses.append(
            [f"J{current_index}", analysts_consensus.strip() or unknown_consensus]
        )
        current_index += 1
        time.sleep(randint(3, 10))
    logging.info("...finished scraping consensuses.")
    driver.close()

    add_consensus_column(file_with_predictions, sheet_name, ticker_consensuses)
