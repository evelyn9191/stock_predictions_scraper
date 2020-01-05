import logging
from datetime import date

import openpyxl as openpyxl

from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

from scrape_consensus import add_consensus
from scrape_predictions import check_latest_predictions, get_stock_predictions, write_to_excel

logging.getLogger().setLevel(logging.INFO)


def apply_formatting(output_file: str, sheetname: str) -> None:
    wb = openpyxl.load_workbook(output_file)
    ws = wb[sheetname]

    # Change size of the column with company names and wrap the text
    ws.column_dimensions["A"].width = 25
    for cell in ws["A"]:
        cell.alignment = Alignment(wrap_text=True)

    # Add column filters
    ws.auto_filter.ref = ws.dimensions

    # Change format of predictions from string to percentage
    percentage_columns = ["E", "F", "G", "H", "I"]
    for col in percentage_columns:
        for cell in ws[col]:
            cell.number_format = FORMAT_PERCENTAGE_00

    # Add conditional formatting for predictions percentages
    green_background = PatternFill(bgColor="C6EFCE")
    diff_style = DifferentialStyle(fill=green_background)
    rule = Rule(type="cellIs", operator="greaterThan", dxf=diff_style, formula=["30.00%"])
    ws.conditional_formatting.add(f"F2:G{ws.max_row}", rule)

    # Hide the column with 5 year predictions
    ws.column_dimensions["I"].hidden = True

    wb.save(output_file)


def run_scraper():
    logging.info("Running scraper...")

    file_with_stocks = "data/revolut_stocks.xlsx"
    file_with_predictions = "data/predictions.xlsx"
    sheet_name = date.today().strftime("%d-%m-%Y")

    old_predictions = check_latest_predictions(file_with_predictions)

    if old_predictions:
        stock_predictions = get_stock_predictions(file_with_stocks)
        write_to_excel(file_with_stocks, file_with_predictions, stock_predictions, sheet_name)
        add_consensus(file_with_predictions, sheet_name)
        apply_formatting(file_with_predictions, sheet_name)

    logging.info("...scraping finished")


if __name__ == "__main__":
    run_scraper()
